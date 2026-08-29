"""Lớp dịch vụ: gửi Zalo, lưu file, tính toạ độ, tính KPI, xuất Excel.

Phần tích hợp OpenAI (gợi ý mô tả, trợ lý AI hỏi-đáp) đã tách sang
dich_vu_ai.py — xem file đó nếu cần sửa/đọc phần AI.
"""
import math
import os
import re
import uuid
from datetime import datetime, date, timedelta
from io import BytesIO

import requests
from flask import current_app
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image as PILImage
from werkzeug.utils import secure_filename

from extensions import db
from models import (BuoiNghi, ChamCong, CongViec, DanhGia, DiemChamCong, DoUuTien,
                    GhiChuNop, LogZalo, LoaiDinhKem, MucSao, NguoiDung, TrangThai,
                    VaiTro, gio_vn_hien_tai, ngay_vn_hien_tai)

# ---------------------------------------------------------------------------
# GỬI ZALO
# ---------------------------------------------------------------------------

def _bot_token(nguoi_dung: NguoiDung | None) -> str:
    if nguoi_dung and nguoi_dung.bot_zalo and nguoi_dung.bot_zalo.dang_hoat_dong:
        return nguoi_dung.bot_zalo.token
    return current_app.config["ZALO_DEFAULT_BOT_TOKEN"]


def gui_zalo(chat_id: str, noi_dung: str, nguoi_dung: NguoiDung | None = None,
             cong_viec: CongViec | None = None, token_ghi_de: str | None = None) -> bool:
    """Gửi 1 tin nhắn text tới 1 chat Zalo. Luôn ghi log, không bao giờ raise.

    Không để lỗi Zalo làm hỏng giao dịch chính (giao việc, duyệt việc...).
    token_ghi_de dùng khi gửi không gắn với 1 nhân viên cụ thể (VD: nhóm QL).
    """
    log = LogZalo(
        nguoi_dung_id=nguoi_dung.id if nguoi_dung else None,
        cong_viec_id=cong_viec.id if cong_viec else None,
        chat_id=chat_id,
        noi_dung=noi_dung,
    )
    token = token_ghi_de or _bot_token(nguoi_dung)
    if not chat_id or not token:
        log.phan_hoi = "Thiếu chat_id hoặc bot token"
        db.session.add(log)
        return False

    url = f"{current_app.config['ZALO_API_BASE']}{token}/sendMessage"
    try:
        r = requests.post(url, json={"chat_id": chat_id, "text": noi_dung}, timeout=10)
        log.thanh_cong = r.ok
        log.phan_hoi = r.text[:1000]
    except Exception as e:  # noqa: BLE001
        log.thanh_cong = False
        log.phan_hoi = f"{type(e).__name__}: {e}"[:1000]
    db.session.add(log)
    return log.thanh_cong


def gui_cho_nhan_vien(nv: NguoiDung, noi_dung: str, cong_viec: CongViec | None = None) -> bool:
    return gui_zalo(nv.zalo_group_id, noi_dung, nguoi_dung=nv, cong_viec=cong_viec)


def gui_nhom_ql(noi_dung: str, cong_viec: CongViec | None = None) -> bool:
    """Gửi vào nhóm Quản Lý. Dùng token riêng của bot phụ trách nhóm QL."""
    token = current_app.config["ZALO_BOT_TOKEN_QL"] or current_app.config["ZALO_DEFAULT_BOT_TOKEN"]
    return gui_zalo(current_app.config["ZALO_GROUP_QL"], noi_dung,
                    cong_viec=cong_viec, token_ghi_de=token)


# ---------------------------------------------------------------------------
# NGHỈ PHÉP
# ---------------------------------------------------------------------------

def co_nghi_phep_buoi(nguoi_dung_id: int, ngay, buoi: str) -> bool:
    """True nếu nhân viên có nghỉ phép đã ghi nhận cho đúng buổi này (hoặc
    cả ngày) — dùng để bỏ qua tính đi trễ/về sớm cho đúng buổi đã xin nghỉ."""
    from models import XinNghi
    return XinNghi.query.filter(
        XinNghi.nguoi_dung_id == nguoi_dung_id,
        XinNghi.ngay == ngay,
        XinNghi.buoi.in_((buoi, BuoiNghi.CA_NGAY)),
    ).first() is not None


def bao_xin_nghi(nguoi_dung: NguoiDung, ngay_dau, ngay_cuoi, buoi: str,
                  ly_do: str, so_ngay_cong: float, ban_giao_cho: NguoiDung | None,
                  duong_dan_pdf: str):
    """Báo 1 tin gộp vào nhóm QL khi có người xin nghỉ (đã tự động duyệt qua
    đơn PDF ký điện tử) — không cần ai duyệt tay, nhưng vẫn cho sếp biết."""
    if ngay_dau == ngay_cuoi:
        thoi_gian = f"Ngày {ngay_dau:%d/%m/%Y} — {BuoiNghi.NHAN.get(buoi, buoi)}"
    else:
        thoi_gian = f"Từ ngày {ngay_dau:%d/%m/%Y} đến ngày {ngay_cuoi:%d/%m/%Y} (cả ngày)"
    nd = (
        f"🏖️ Đã ghi nhận nghỉ phép — tự động duyệt qua đơn ký điện tử\n\n"
        f"{nguoi_dung.ho_ten} ({nguoi_dung.ma_dinh_danh})\n"
        f"{thoi_gian}\n"
        f"Tính {so_ngay_cong:g} ngày công.\n"
        f"Lý do: {ly_do}"
    )
    if ban_giao_cho:
        nd += f"\nBàn giao công việc cho: {ban_giao_cho.ho_ten}"
    nd += f"\n\nXem đơn (PDF):\n{current_app.config['BASE_URL']}/media/{duong_dan_pdf}"
    gui_nhom_ql(nd)


def _dang_ky_font_unicode():
    """Đăng ký font DejaVu Sans (có dấu tiếng Việt) cho reportlab — chỉ cần
    làm 1 lần. Font mặc định của reportlab (Helvetica...) không có dấu
    tiếng Việt nên không dùng được cho đơn xin nghỉ."""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    if "DejaVu" in pdfmetrics.getRegisteredFontNames():
        return
    thu_muc_font = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "fonts")
    pdfmetrics.registerFont(TTFont("DejaVu", os.path.join(thu_muc_font, "DejaVuSans.ttf")))
    pdfmetrics.registerFont(TTFont("DejaVu-Bold", os.path.join(thu_muc_font, "DejaVuSans-Bold.ttf")))


def cat_anh_chu_ky(chu_ky_png: bytes, le: int = 8) -> bytes:
    """Cắt ảnh chữ ký (canvas PNG nền trong suốt) về đúng vùng có mực, chừa
    thêm 1 chút lề — dùng CHUNG cho cả lúc xem trước (route xem_truoc_chu_ky)
    và lúc nhúng thật vào PDF (tao_pdf_don_xin_nghi), để "xem trước" luôn
    khớp 100% với PDF thật, không lệch thuật toán giữa 2 nơi.

    Dùng PIL.Image.getbbox() (toàn bộ pixel có mực, dù đậm hay nhạt) — từng
    thử cách "chỉ tính cột/dòng có mật độ mực đủ dày" để loại đuôi ký mảnh,
    nhưng cách đó dễ cắt nhầm luôn cả nét chữ ký thật (không chỉ đuôi) nên
    bỏ, chấp nhận đổi lại: nếu ai ký kèm 1 nét vuốt dài tách hẳn ra xa thân
    chữ ký, bbox sẽ rộng hơn 1 chút — vẫn tốt hơn nhiều so với mất nét.

    Trả về bytes PNG đã cắt (nguyên bản nếu canvas trắng, không có gì để cắt).
    """
    anh = PILImage.open(BytesIO(chu_ky_png))
    hop = anh.getbbox()
    if hop:
        x0, y0, x1, y1 = hop
        x0 = max(0, x0 - le)
        y0 = max(0, y0 - le)
        x1 = min(anh.width, x1 + le)
        y1 = min(anh.height, y1 + le)
        anh = anh.crop((x0, y0, x1, y1))
    buf = BytesIO()
    anh.save(buf, format="PNG")
    return buf.getvalue()


def tao_pdf_don_xin_nghi(nguoi_dung: NguoiDung, ngay_dau: date, ngay_cuoi: date,
                          buoi: str, ly_do: str, ban_giao_cho: NguoiDung | None,
                          chu_ky_png: bytes) -> bytes:
    """Sinh file PDF đơn xin nghỉ phép theo đúng thể thức văn bản hành chính
    (quốc hiệu tiêu ngữ 2 cột, số văn bản, bảng thông tin người làm đơn),
    đã điền sẵn thông tin nhân viên + chữ ký điện tử vừa vẽ — thay cho việc
    nhân viên phải in ra, ký tay, chụp ảnh rồi tải lên như trước. Trả về
    nội dung PDF dạng bytes."""
    from reportlab.lib.colors import HexColor
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_RIGHT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import Image as RLImage
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    from reportlab.platypus.flowables import HRFlowable

    _dang_ky_font_unicode()

    VANG_DAM = HexColor("#8A6A1E")
    XAM = HexColor("#75746E")
    XAM_VIEN = HexColor("#E8E6E1")
    DEN = HexColor("#1C1C1A")
    XANH = HexColor("#3F7D55")
    XANH_NHAT = HexColor("#EAF3EC")

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=1.4 * cm, bottomMargin=1.4 * cm, leftMargin=2.3 * cm, rightMargin=2.3 * cm,
    )
    rong_trang = A4[0] - 2.3 * cm - 2.3 * cm

    kieu_quoc_hieu = ParagraphStyle("quoc_hieu", fontName="DejaVu-Bold", fontSize=13, alignment=TA_CENTER, leading=16)
    kieu_tieu_ngu = ParagraphStyle("tieu_ngu", fontName="DejaVu", fontSize=11.5, alignment=TA_CENTER, leading=15)
    kieu_tieu_de = ParagraphStyle("tieu_de", fontName="DejaVu-Bold", fontSize=16, alignment=TA_CENTER,
                                  textColor=DEN, spaceBefore=8, spaceAfter=14)
    kieu_de_muc = ParagraphStyle("de_muc", fontName="DejaVu-Bold", fontSize=10.5, textColor=VANG_DAM,
                                 spaceBefore=12, spaceAfter=6)
    kieu_thuong = ParagraphStyle("thuong", fontName="DejaVu", fontSize=10.5, alignment=TA_JUSTIFY,
                                 leading=15, spaceAfter=6, textColor=DEN)
    kieu_kinh_gui_nhan = ParagraphStyle("kinh_gui_nhan", fontName="DejaVu-Bold", fontSize=11, textColor=DEN)
    kieu_kinh_gui_gt = ParagraphStyle("kinh_gui_gt", fontName="DejaVu", fontSize=11, textColor=DEN,
                                      leading=15, spaceAfter=2)
    kieu_nhan_tt = ParagraphStyle("nhan_tt", fontName="DejaVu-Bold", fontSize=10, textColor=XAM)
    kieu_gt_tt = ParagraphStyle("gt_tt", fontName="DejaVu", fontSize=11, textColor=DEN)
    kieu_phai = ParagraphStyle("phai", fontName="DejaVu", fontSize=10.5, alignment=TA_RIGHT, textColor=DEN)
    kieu_phai_dam = ParagraphStyle("phai_dam", fontName="DejaVu-Bold", fontSize=11, alignment=TA_RIGHT, textColor=DEN)
    kieu_phai_nho = ParagraphStyle("phai_nho", fontName="DejaVu", fontSize=9, alignment=TA_RIGHT, textColor=XAM)
    kieu_xn_tieu_de = ParagraphStyle("xn_tieu_de", fontName="DejaVu-Bold", fontSize=9, alignment=TA_CENTER, textColor=XANH)
    kieu_xn_nd = ParagraphStyle("xn_nd", fontName="DejaVu", fontSize=8.7, alignment=TA_CENTER, textColor=DEN, leading=12.5)
    kieu_xn_tg = ParagraphStyle("xn_tg", fontName="DejaVu", fontSize=8, alignment=TA_CENTER, textColor=XAM, spaceBefore=4)

    bay_gio = gio_vn_hien_tai()

    # Quốc hiệu/tiêu ngữ đứng ngay đầu trang, căn giữa, chiếm trọn chiều
    # ngang trang — đúng thể thức văn bản hành chính, không kèm theo bất
    # kỳ letterhead công ty nào ở trên nó.
    noi_dung = [
        Paragraph("CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM", kieu_quoc_hieu),
        Paragraph("Độc lập - Tự do - Hạnh phúc", kieu_tieu_ngu),
        HRFlowable(width="32%", thickness=1, color=DEN, hAlign="CENTER", spaceBefore=3, spaceAfter=6),
        Paragraph("ĐƠN XIN NGHỈ PHÉP", kieu_tieu_de),
    ]

    # "Kính gửi" nhiều nơi nhận thì mỗi nơi xuống 1 dòng riêng, thẳng hàng
    # dưới dòng đầu — không viết liền thành 1 câu nối bằng dấu phẩy.
    cac_noi_nhan = ["Ban Giám đốc BRICON"]
    if nguoi_dung.vai_tro == VaiTro.NHAN_VIEN and nguoi_dung.bo_phan:
        cac_noi_nhan.append(f"Quản lý bộ phận {nguoi_dung.bo_phan.ten}")
    bang_kinh_gui = Table(
        [[Paragraph("Kính gửi:", kieu_kinh_gui_nhan),
          [Paragraph(dong, kieu_kinh_gui_gt) for dong in cac_noi_nhan]]],
        colWidths=[2.6 * cm, rong_trang - 2.6 * cm],
    )
    bang_kinh_gui.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    noi_dung.append(bang_kinh_gui)

    # ------------------------------------------------ bảng thông tin người làm đơn
    noi_dung.append(Paragraph("THÔNG TIN NGƯỜI LÀM ĐƠN", kieu_de_muc))

    def _hang_tt(nhan, gt):
        return [Paragraph(nhan, kieu_nhan_tt), Paragraph(gt, kieu_gt_tt)]

    du_lieu_tt = [
        _hang_tt("Họ và tên", nguoi_dung.ho_ten),
        _hang_tt("Mã nhân viên", nguoi_dung.ma_dinh_danh),
        _hang_tt("Chức vụ", nguoi_dung.chuc_vu.ten if nguoi_dung.chuc_vu else "—"),
        _hang_tt("Bộ phận", nguoi_dung.bo_phan.ten if nguoi_dung.bo_phan else "—"),
        _hang_tt("Điện thoại liên hệ", nguoi_dung.so_dien_thoai),
    ]
    bang_tt = Table(du_lieu_tt, colWidths=[rong_trang * 0.28, rong_trang * 0.72])
    bang_tt.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LINEBELOW", (0, 0), (-1, -2), 0.5, XAM_VIEN),
        ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0),
    ]))
    noi_dung.append(bang_tt)

    # ------------------------------------------------ nội dung xin nghỉ
    noi_dung.append(Paragraph("NỘI DUNG XIN NGHỈ", kieu_de_muc))

    if ngay_dau == ngay_cuoi:
        so_ngay_hien = f"{BuoiNghi.SO_NGAY.get(buoi, 1.0):g} ngày ({BuoiNghi.NHAN.get(buoi, '')})"
        khoang_ngay = f"ngày {ngay_dau:%d/%m/%Y}"
    else:
        so_ngay = (ngay_cuoi - ngay_dau).days + 1
        so_ngay_hien = f"{so_ngay} ngày"
        khoang_ngay = f"từ ngày {ngay_dau:%d/%m/%Y} đến ngày {ngay_cuoi:%d/%m/%Y}"
    noi_dung.append(Paragraph(
        f"Kính đề Ban Giám đốc cho tôi nghỉ phép {so_ngay_hien}, {khoang_ngay}.", kieu_thuong))
    noi_dung.append(Paragraph(f"<b>Lý do:</b> {ly_do}", kieu_thuong))

    if ban_giao_cho:
        chuc_vu_bgc = f" ({ban_giao_cho.chuc_vu.ten})" if ban_giao_cho.chuc_vu else ""
        noi_dung.append(Paragraph(
            f"Tôi đã bàn giao công việc trong thời gian nghỉ phép lại cho: "
            f"<b>{ban_giao_cho.ho_ten}</b>{chuc_vu_bgc}.", kieu_thuong))

    noi_dung.append(Paragraph(
        "Tôi xin cam kết thực hiện nghiêm túc thời gian nghỉ phép như đã đăng ký ở trên, "
        "chủ động sắp xếp và bàn giao công việc đầy đủ trước khi nghỉ để không làm ảnh hưởng "
        "đến tiến độ chung của bộ phận và công ty. Trong trường hợp cần thiết, tôi vẫn sẵn sàng "
        "phối hợp xử lý công việc từ xa. Tôi cam kết quay trở lại làm việc đúng thời hạn đã đăng ký "
        "và xin hoàn toàn chịu trách nhiệm nếu để xảy ra sai sót do việc nghỉ phép của mình gây ra.",
        kieu_thuong,
    ))
    noi_dung.append(Paragraph(
        "Kính mong Ban Giám đốc xem xét và tạo điều kiện chấp thuận cho tôi được nghỉ phép "
        "theo nguyện vọng trên. Tôi xin trân trọng cảm ơn!", kieu_thuong,
    ))
    noi_dung.append(Spacer(1, 8))

    # ------------------------------------------------ khung xác nhận + ký tên
    # Cắt bằng đúng hàm cat_anh_chu_ky() dùng chung với route xem trước, để
    # "xem trước" trên form luôn khớp 100% với PDF thật.
    buf_anh_ky = BytesIO(cat_anh_chu_ky(chu_ky_png))

    anh_ky = RLImage(buf_anh_ky)
    ti_le = min(1.0, (5.5 * cm) / anh_ky.imageWidth)
    anh_ky.drawWidth = anh_ky.imageWidth * ti_le
    anh_ky.drawHeight = anh_ky.imageHeight * ti_le

    # QUAN TRỌNG: Image.hAlign KHÔNG căn đúng mép phải thật của cột khi ảnh
    # nằm trong 1 list nhiều phần tử lồng bên trong ô của Table (đã đo trực
    # tiếp trên PDF thật bằng PyMuPDF — lệch trái cố định dù chữ ký gì đi
    # nữa, không liên quan gì tới bước cắt ảnh ở trên). Sửa bằng cách bọc
    # ảnh vào 1 bảng con riêng, dùng ALIGN cấp ô (đáng tin cậy hơn hẳn
    # Image.hAlign trong trường hợp lồng nhau này) để căn đúng mép phải
    # thật của cột "o_ky_ten" bên dưới.
    rong_cot_ky = rong_trang * 0.62 - 14  # trừ đúng LEFTPADDING đã đặt cho cột này
    bang_anh_ky = Table([[anh_ky]], colWidths=[rong_cot_ky])
    bang_anh_ky.setStyle(TableStyle([
        ("ALIGN", (0, 0), (0, 0), "RIGHT"),
        ("VALIGN", (0, 0), (0, 0), "TOP"),
        ("LEFTPADDING", (0, 0), (0, 0), 0), ("RIGHTPADDING", (0, 0), (0, 0), 0),
        ("TOPPADDING", (0, 0), (0, 0), 0), ("BOTTOMPADDING", (0, 0), (0, 0), 0),
    ]))

    o_xac_nhan = [
        Spacer(1, 8),
        Paragraph("✓ HỆ THỐNG TỰ ĐỘNG PHÊ DUYỆT", kieu_xn_tieu_de),
        Spacer(1, 5),
        Paragraph("Đơn được ghi nhận và phê duyệt tự động ngay sau khi người làm đơn ký điện tử.", kieu_xn_nd),
        Paragraph(f"Lúc {bay_gio:%H:%M} ngày {bay_gio:%d/%m/%Y}", kieu_xn_tg),
    ]
    o_ky_ten = [
        Paragraph(f"Tp. Hồ Chí Minh, ngày {bay_gio:%d} tháng {bay_gio:%m} năm {bay_gio:%Y}", kieu_phai),
        Paragraph("NGƯỜI LÀM ĐƠN", kieu_phai_dam),
        Paragraph("(Đã ký điện tử)", kieu_phai_nho),
        Spacer(1, 6),
        bang_anh_ky,
        Paragraph(nguoi_dung.ho_ten, kieu_phai_dam),
    ]
    bang_ky = Table([[o_xac_nhan, o_ky_ten]], colWidths=[rong_trang * 0.38, rong_trang * 0.62])
    bang_ky.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOX", (0, 0), (0, 0), 0.75, XANH),
        ("BACKGROUND", (0, 0), (0, 0), XANH_NHAT),
        ("TOPPADDING", (0, 0), (0, 0), 4), ("BOTTOMPADDING", (0, 0), (0, 0), 10),
        ("LEFTPADDING", (0, 0), (0, 0), 10), ("RIGHTPADDING", (0, 0), (0, 0), 10),
        ("LEFTPADDING", (1, 0), (1, 0), 14), ("RIGHTPADDING", (1, 0), (1, 0), 0),
        ("TOPPADDING", (1, 0), (1, 0), 0), ("BOTTOMPADDING", (1, 0), (1, 0), 0),
    ]))
    noi_dung.append(bang_ky)

    doc.build(noi_dung)
    return buf.getvalue()


def luu_pdf_don_xin_nghi(pdf_bytes: bytes) -> str:
    """Lưu PDF đơn xin nghỉ vừa sinh vào ổ đĩa, trả về đường dẫn tương đối
    (giống quy ước luu_file — dùng chung thư mục 'xin-nghi' như ảnh cũ)."""
    tuong_doi, tuyet_doi = _duong_dan_moi("don-xin-nghi.pdf", "xin-nghi")
    with open(tuyet_doi, "wb") as f:
        f.write(pdf_bytes)
    return tuong_doi


def lay_cac_chat_gan_day(token: str) -> tuple[list[dict], str | None, str]:
    """Gọi getUpdates để tìm các nhóm/chat bot vừa nhận được tin nhắn — dùng
    để dò chat_id của 1 nhóm mới thêm bot vào.

    Trả về (danh sách chat đã nhận diện được, thông báo lỗi nếu có, JSON thô
    để đối chiếu tay nếu cách đọc tự động ở đây thiếu do khác định dạng).
    """
    url = f"{current_app.config['ZALO_API_BASE']}{token}/getUpdates"
    try:
        r = requests.get(url, timeout=10)
        tho = r.text
        du_lieu = r.json()
    except Exception as e:  # noqa: BLE001
        return [], f"Không gọi được API: {type(e).__name__}: {e}", ""

    if du_lieu.get("ok") is False:
        return [], du_lieu.get("description") or "Zalo báo lỗi nhưng không rõ nguyên nhân.", tho

    ds = du_lieu.get("result")
    if ds is None:
        ds = du_lieu.get("updates") or []
    if isinstance(ds, dict):
        ds = [ds]  # Zalo trả 1 object đơn (không phải mảng) khi chỉ có 1 cập nhật
    if not isinstance(ds, list):
        return [], "Zalo trả về dữ liệu không đúng định dạng mong đợi — xem JSON thô bên dưới.", tho

    thay: dict = {}
    for cap_nhat in ds:
        tin = cap_nhat.get("message") or {}
        chat = tin.get("chat") or {}
        nguoi_gui = tin.get("from") or {}
        chat_id = chat.get("id")
        if not chat_id:
            continue
        thay[chat_id] = {
            "chat_id": chat_id,
            "loai": chat.get("chat_type") or chat.get("type") or "—",
            "ten": (chat.get("title") or chat.get("name")
                   or f"(Zalo không trả tên nhóm — tin gần nhất từ {nguoi_gui.get('display_name', '?')})"),
        }
    return list(thay.values()), None, tho


# ---------------------------------------------------------------------------
# WEBHOOK ZALO — thay thế n8n cho các lệnh chat trực tiếp trong nhóm
# ---------------------------------------------------------------------------

def dat_webhook(bot: "BotZalo") -> tuple[bool, str]:
    """Đăng ký webhook với Zalo cho 1 bot — Zalo sẽ tự POST tin nhắn mới về
    /webhook/zalo/<bot_id> thay vì phải gọi getUpdates để hỏi."""
    base = current_app.config["BASE_URL"]
    if not base.startswith("https://"):
        return False, ("BASE_URL hiện không phải https — Zalo sẽ từ chối webhook. "
                       "Cần deploy với domain thật + SSL trước khi đặt webhook.")

    url_webhook = f"{base}/webhook/zalo/{bot.id}"
    api_url = f"{current_app.config['ZALO_API_BASE']}{bot.token}/setWebhook"
    try:
        r = requests.post(api_url, json={
            "url": url_webhook,
            "secret_token": bot.webhook_secret,
        }, timeout=10)
        return r.ok, r.text
    except Exception as e:  # noqa: BLE001
        return False, f"{type(e).__name__}: {e}"


def xoa_webhook(bot: "BotZalo") -> tuple[bool, str]:
    api_url = f"{current_app.config['ZALO_API_BASE']}{bot.token}/deleteWebhook"
    try:
        r = requests.post(api_url, timeout=10)
        return r.ok, r.text
    except Exception as e:  # noqa: BLE001
        return False, f"{type(e).__name__}: {e}"


def xu_ly_webhook_zalo(bot: "BotZalo", du_lieu: dict):
    """Xử lý 1 sự kiện webhook Zalo gửi tới.
    - "/id": trả zalo_group_id + zalo_user_id (dò nhóm khi thiết lập).
    - Tag bot + có chữ "giao việc": trả link trang Giao việc mới.
    - Tag bot + có chữ "kpi": trả link trang KPI.
    - Tag bot, còn lại: trả link trang chủ hệ thống.

    Zalo không công bố rõ trường "đã tag bot" trong webhook, nên nhận diện
    bằng nội dung tin nhắn: có tên bot hoặc dấu "@" thì coi là đã tag —
    tránh trả lời tràn lan mọi tin nhắn thường trong nhóm. Muốn thêm lệnh
    khác thì thêm nhánh if bên dưới, cùng một nơi.
    """
    tin = du_lieu.get("message") or {}
    chat = tin.get("chat") or {}
    nguoi_gui = tin.get("from") or {}
    text = (tin.get("text") or "").strip()

    chat_id = chat.get("id")
    if not chat_id or not text:
        return

    text_thuong = text.lower()

    if text_thuong.startswith("/id"):
        nd = (
            f"🆔 Thông tin nhóm/chat này\n\n"
            f"zalo_group_id: {chat_id}\n"
            f"zalo_user_id: {nguoi_gui.get('id', '—')}"
        )
        gui_zalo(chat_id, nd, token_ghi_de=bot.token)
        return

    da_tag = "@" in text or (bot.ten and bot.ten.lower() in text_thuong)
    if not da_tag:
        return

    base = current_app.config["BASE_URL"]
    if "giao việc" in text_thuong or "giao viec" in text_thuong:
        nd = f"📌 Vào đây để giao việc mới:\n{base}/viec/moi"
    elif "kpi" in text_thuong:
        nd = f"📊 Vào đây để xem KPI:\n{base}/kpi"
    else:
        nd = f"👋 Vào hệ thống BRICON WORK tại đây:\n{base}/"
    gui_zalo(chat_id, nd, token_ghi_de=bot.token)


# ---- Nội dung tin nhắn ------------------------------------------------------

def _han_str(viec: CongViec) -> str:
    return viec.han.strftime("%H:%M %d/%m/%Y") if viec.han else "không đặt hạn"


def bao_giao_viec(viec: CongViec):
    nd = (
        f"📌 Bạn có 1 công việc mới: {viec.ten_uu_tien}\n\n"
        f"[{viec.ma}] {viec.tieu_de}\n"
        f"Người giao: {viec.nguoi_giao.ho_ten}\n"
        f"Hạn: {_han_str(viec)}\n\n"
        f"Bấm vào xem chi tiết và gửi đối chứng:\n{viec.link}"
    )
    gui_cho_nhan_vien(viec.nguoi_nhan, nd, viec)


def bao_giao_viec_hang_ngay(nguoi_nhan: NguoiDung, tieu_de: str, ngay_dau, ngay_cuoi,
                             gio_han: str, so_luong: int):
    """Báo gộp 1 tin duy nhất khi giao 1 loạt việc Hằng ngày cho 1 nhân viên,
    thay vì báo riêng từng ngày (mỗi ngày là 1 công việc riêng trong hệ thống,
    nhưng chỉ cần 1 tin Zalo tóm tắt)."""
    if ngay_dau == ngay_cuoi:
        thoi_gian = f"Vào ngày {ngay_dau:%d/%m/%Y}, hạn lúc {gio_han}"
    else:
        thoi_gian = f"Từ ngày {ngay_dau:%d/%m/%Y} tới ngày {ngay_cuoi:%d/%m/%Y}, hạn lúc {gio_han} mỗi ngày"
    nd = (
        f"📅 Bạn có việc hằng ngày mới: {tieu_de}\n\n"
        f"{thoi_gian}.\n"
        f"Tổng cộng {so_luong} công việc đã được tạo.\n\n"
        f"Xem danh sách công việc:\n{current_app.config['BASE_URL']}/viec"
    )
    gui_cho_nhan_vien(nguoi_nhan, nd)


def bao_doi_uu_tien(viec: CongViec, uu_tien_cu_ten: str, nguoi_doi: NguoiDung):
    """Báo cho nhân viên khi Sếp/Admin kéo thả đổi mức độ công việc trên
    trang Công việc."""
    nd = (
        f"🔄 Mức độ công việc đã được đổi\n\n"
        f"[{viec.ma}] {viec.tieu_de}\n"
        f"Từ: {uu_tien_cu_ten} → Sang: {viec.ten_uu_tien}\n"
        f"Người đổi: {nguoi_doi.ho_ten}\n\n"
        f"Xem chi tiết:\n{viec.link}"
    )
    gui_cho_nhan_vien(viec.nguoi_nhan, nd, viec)


def bao_dat_lai_han(viec: CongViec, han_cu, nguoi_doi: NguoiDung):
    """Báo cho nhân viên khi Sếp/Admin đặt lại (hoặc đặt lần đầu) hạn hoàn
    thành cho 1 công việc."""
    han_cu_hien = han_cu.strftime("%H:%M %d/%m/%Y") if han_cu else "(chưa đặt hạn)"
    han_moi_hien = viec.han.strftime("%H:%M %d/%m/%Y") if viec.han else "(đã bỏ hạn)"
    nd = (
        f"📅 Hạn hoàn thành đã được cập nhật\n\n"
        f"[{viec.ma}] {viec.tieu_de}\n"
        f"Từ: {han_cu_hien} → Sang: {han_moi_hien}\n"
        f"Người đổi: {nguoi_doi.ho_ten}\n\n"
        f"Xem chi tiết:\n{viec.link}"
    )
    gui_cho_nhan_vien(viec.nguoi_nhan, nd, viec)


# ---------------------------------------------------------------------------
# BÁO CÁO ZALO THEO LỊCH TRONG NGÀY (chạy bằng cron, xem app.py)
# ---------------------------------------------------------------------------

def dieu_kien_viec_trong_ngay(ngay: date):
    """Điều kiện lọc SQLAlchemy dùng chung cho Dashboard + Trợ lý AI: việc
    có hạn rơi đúng 'ngay' (như trước giờ), HOẶC có đặt "ngày bắt đầu" và
    'ngay' nằm trong khoảng [ngay_bat_dau, ngày của han] — để việc có
    khoảng thời gian dài (VD chờ duyệt nhiều ngày) hiện xuyên suốt trong
    "Việc hằng ngày", không chỉ đúng ngày hạn cuối cùng."""
    dau_ngay = datetime.combine(ngay, datetime.min.time())
    cuoi_ngay = datetime.combine(ngay, datetime.max.time())
    return db.or_(
        db.and_(
            CongViec.ngay_bat_dau.is_(None),
            CongViec.han >= dau_ngay,
            CongViec.han <= cuoi_ngay,
        ),
        db.and_(
            CongViec.ngay_bat_dau.isnot(None),
            CongViec.ngay_bat_dau <= ngay,
            CongViec.han >= dau_ngay,
        ),
    )


def _nhan_vien_khong_phai_admin_sep() -> list[NguoiDung]:
    return NguoiDung.query.filter_by(dang_hoat_dong=True).filter(
        NguoiDung.vai_tro.notin_((VaiTro.ADMIN, VaiTro.SEP))
    ).all()


def nhac_viec_sap_qua_han(phut_truoc: int = 30) -> int:
    """Nhắc việc sắp tới hạn (mặc định trong vòng 30 phút tới), CHỈ gửi cho
    đúng nhân viên đang nhận việc đó — không báo nhóm QL. Nên chạy bằng
    cron thường xuyên (VD mỗi 5 phút) để bắt kịp mốc giờ; dùng cột
    da_nhac_sap_qua_han để mỗi việc chỉ nhắc đúng 1 lần, không bị nhắc lặp
    lại liên tục trong suốt khung 30 phút đó. Chỉ nhắc việc CHƯA nộp gì
    (mới/đang làm/phải làm lại) — việc đang chờ duyệt thì đã ngoài tầm tay
    nhân viên, không cần nhắc."""
    bay_gio = gio_vn_hien_tai()
    moc_xa = bay_gio + timedelta(minutes=phut_truoc)
    viecs = CongViec.query.filter(
        CongViec.trang_thai.in_(TrangThai.CHUA_XONG),
        CongViec.han.isnot(None),
        CongViec.han > bay_gio,
        CongViec.han <= moc_xa,
        CongViec.da_nhac_sap_qua_han.is_(False),
    ).all()
    for v in viecs:
        nd = (
            f"⏳ Việc sắp tới hạn (còn dưới {phut_truoc} phút)\n\n"
            f"[{v.ma}] {v.tieu_de}\n"
            f"Hạn: {v.han:%H:%M %d/%m/%Y}\n\n"
            f"Xem chi tiết và gửi đối chứng ngay:\n{v.link}"
        )
        gui_cho_nhan_vien(v.nguoi_nhan, nd, v)
        v.da_nhac_sap_qua_han = True
    return len(viecs)


def nhac_cham_cong_sang() -> int:
    """7h45: gửi link chấm công cho từng nhân viên/quản lý (trừ Sếp/Admin —
    2 role này không tự chấm công). Trả về số người đã gửi."""
    base = current_app.config["BASE_URL"]
    nd = f"⏰ Nhớ chấm công vào nhé!\n\n{base}/cham-cong"
    ds = _nhan_vien_khong_phai_admin_sep()
    for nv in ds:
        gui_cho_nhan_vien(nv, nd)
    return len(ds)


def nhac_cham_cong_chieu() -> int:
    """17h32: nhắc chấm công RA — chỉ nhắc đúng người đã chấm công VÀO hôm
    nay nhưng CHƯA chấm ra (không làm phiền người đã ra rồi hoặc chưa từng
    chấm vào). Trả về số người đã gửi."""
    base = current_app.config["BASE_URL"]
    hom_nay = ngay_vn_hien_tai()
    ds = (ChamCong.query.filter_by(ngay=hom_nay)
          .filter(ChamCong.gio_vao.isnot(None), ChamCong.gio_ra.is_(None))
          .all())
    nd = f"⏰ Nhớ chấm công ra trước khi về nhé!\n\n{base}/cham-cong"
    so_luong = 0
    for cc in ds:
        nv = cc.nguoi_dung
        if not nv or not nv.dang_hoat_dong or nv.vai_tro in (VaiTro.ADMIN, VaiTro.SEP):
            continue
        gui_cho_nhan_vien(nv, nd)
        so_luong += 1
    return so_luong


def nhac_viec_hom_nay() -> int:
    """8h00: gửi link xem việc hôm nay cho từng nhân viên/quản lý (trừ
    Sếp/Admin). Trả về số người đã gửi."""
    base = current_app.config["BASE_URL"]
    nd = f"📋 Xem công việc hôm nay của bạn:\n\n{base}/viec"
    ds = _nhan_vien_khong_phai_admin_sep()
    for nv in ds:
        gui_cho_nhan_vien(nv, nd)
    return len(ds)


def bao_cao_sang_cho_sep():
    """8h10: báo cáo nhanh đầu ngày cho Sếp/Quản lý — gửi vào nhóm QL."""
    from models import XinNghi
    hom_nay = ngay_vn_hien_tai()
    dau_ngay = datetime.combine(hom_nay, datetime.min.time())
    cuoi_ngay = datetime.combine(hom_nay, datetime.max.time())

    co_mat = ChamCong.query.filter_by(ngay=hom_nay).filter(ChamCong.gio_vao.isnot(None)).count()
    di_tre = ChamCong.query.filter_by(ngay=hom_nay, di_tre=True).count()
    xin_nghi_hom_nay = XinNghi.query.filter_by(ngay=hom_nay).count()
    viec_hom_nay = CongViec.query.filter(CongViec.han >= dau_ngay, CongViec.han <= cuoi_ngay).count()
    qua_han = CongViec.query.filter(
        CongViec.han < gio_vn_hien_tai(), CongViec.trang_thai.in_(TrangThai.CHUA_XONG)).count()
    cho_duyet = CongViec.query.filter_by(trang_thai=TrangThai.CHO_DUYET).count()

    nd = (
        f"📋 BRICON – {hom_nay:%d/%m/%Y}\n\n"
        f"🧑‍💼 {co_mat} nhân viên có mặt\n"
        f"🕐 {di_tre} người đi trễ\n"
        f"📝 {xin_nghi_hom_nay} đơn xin nghỉ hôm nay\n"
        f"📅 {viec_hom_nay} công việc hôm nay\n"
        f"🔴 {qua_han} việc quá hạn\n"
        f"🟡 {cho_duyet} việc chờ đánh giá\n\n"
        f"Chúc anh một ngày làm việc hiệu quả! 💪"
    )
    gui_nhom_ql(nd)


def bao_cao_chieu_cho_sep():
    """17h30: báo cáo tóm tắt cuối ngày cho Sếp/Quản lý — gửi vào nhóm QL.

    Không còn đếm "việc quá hạn" nữa — hệ thống đã tự động đóng + chấm 0★
    MỌI việc quá hạn ngay khi phát hiện (before_request mỗi lần mở trang +
    cron mỗi 5 phút), nên tới giờ báo cáo (17h30) gần như không bao giờ
    còn việc nào thực sự ở trạng thái "quá hạn chưa xử lý" — con số đó chỉ
    gây hiểu lầm. Thay bằng liệt kê từng việc đã bị 0★ hôm nay (tức đã bị
    hệ thống tự đóng do không nộp đúng hạn) kèm link, và liệt kê luôn từng
    việc đang chờ duyệt kèm link để sếp bấm vào duyệt ngay từ Zalo.
    """
    from models import XinNghi
    hom_nay = ngay_vn_hien_tai()
    dau_ngay = datetime.combine(hom_nay, datetime.min.time())
    cuoi_ngay = datetime.combine(hom_nay, datetime.max.time())

    co_mat = ChamCong.query.filter_by(ngay=hom_nay).filter(ChamCong.gio_vao.isnot(None)).count()
    di_tre = ChamCong.query.filter_by(ngay=hom_nay, di_tre=True).count()
    ve_som = ChamCong.query.filter_by(ngay=hom_nay, ve_som=True).count()
    xin_nghi_hom_nay = XinNghi.query.filter_by(ngay=hom_nay).count()
    viec_hom_nay = CongViec.query.filter(CongViec.han >= dau_ngay, CongViec.han <= cuoi_ngay).count()
    hoan_thanh_hom_nay = CongViec.query.filter(
        CongViec.trang_thai == TrangThai.HOAN_THANH,
        CongViec.hoan_thanh_luc >= dau_ngay, CongViec.hoan_thanh_luc <= cuoi_ngay,
    ).count()

    khong_dung_han = CongViec.query.filter(
        CongViec.so_sao_cuoi == 0,
        CongViec.hoan_thanh_luc >= dau_ngay, CongViec.hoan_thanh_luc <= cuoi_ngay,
    ).order_by(CongViec.hoan_thanh_luc).all()
    cho_duyet_ds = CongViec.query.filter_by(
        trang_thai=TrangThai.CHO_DUYET
    ).order_by(CongViec.gui_doi_chung_luc).all()

    nd = (
        f"📊 Tóm tắt cuối ngày – BRICON {hom_nay:%d/%m/%Y}\n\n"
        f"🧑‍💼 {co_mat} nhân viên có mặt hôm nay\n"
        f"🕐 {di_tre} người đi trễ · 🏃 {ve_som} người về sớm\n"
        f"📝 {xin_nghi_hom_nay} đơn xin nghỉ hôm nay\n"
        f"📅 {viec_hom_nay} công việc hôm nay · ✅ {hoan_thanh_hom_nay} đã hoàn thành\n\n"
    )

    if khong_dung_han:
        nd += f"🔴 {len(khong_dung_han)} việc không hoàn thành đúng hạn — 0★:\n" + "\n".join(
            f"• [{v.ma}] {v.tieu_de} — {v.nguoi_nhan.ho_ten}\n  {v.link}"
            for v in khong_dung_han
        )
    else:
        nd += "🔴 Không có việc nào bị 0★ hôm nay."
    nd += "\n\n"

    if cho_duyet_ds:
        nd += f"🟡 {len(cho_duyet_ds)} việc đang chờ duyệt:\n" + "\n".join(
            f"• [{v.ma}] {v.tieu_de} — {v.nguoi_nhan.ho_ten}\n  {v.link}"
            for v in cho_duyet_ds
        )
    else:
        nd += "🟡 Không có việc nào đang chờ duyệt."

    nd += "\n\nMột ngày làm việc nữa đã hoàn tất!"
    gui_nhom_ql(nd)


def bao_cao_thieu_sot():
    """18h00: báo cáo nhân viên nào còn việc chưa nộp/còn thiếu trong ngày
    — gửi vào nhóm QL, liệt kê theo từng người."""
    hom_nay = ngay_vn_hien_tai()
    dau_ngay = datetime.combine(hom_nay, datetime.min.time())
    cuoi_ngay = datetime.combine(hom_nay, datetime.max.time())

    viec_con_thieu = CongViec.query.filter(
        CongViec.han >= dau_ngay, CongViec.han <= cuoi_ngay,
        CongViec.trang_thai.in_(TrangThai.CHUA_XONG),
    ).order_by(CongViec.nguoi_nhan_id).all()

    if not viec_con_thieu:
        gui_nhom_ql(
            f"✅ Hết ngày {hom_nay:%d/%m/%Y}, không còn việc nào tồn đọng "
            f"chưa nộp. Cả đội làm tốt lắm!"
        )
        return

    theo_nguoi: dict[str, list[str]] = {}
    for v in viec_con_thieu:
        theo_nguoi.setdefault(v.nguoi_nhan.ho_ten, []).append(
            f"[{v.ma}] {v.tieu_de} ({v.ten_trang_thai})")

    dong = "\n".join(f"• {ten}: " + "; ".join(vs) for ten, vs in theo_nguoi.items())
    nd = (
        f"📝 Việc còn thiếu cuối ngày {hom_nay:%d/%m/%Y}\n\n"
        f"{dong}\n\n"
        f"Nhắc các bạn hoàn thành và gửi đối chứng sớm nhé."
    )
    gui_nhom_ql(nd)


# ---------------------------------------------------------------------------
# BẢN TIN CÁ NHÂN — gửi riêng cho từng nhân viên lúc 08h00 và 17h30
# ---------------------------------------------------------------------------
# Ghi chú quan trọng: hệ thống KHÔNG bao giờ để việc "quá hạn" tồn đọng —
# mọi việc quá hạn hiệu lực mà chưa nộp gì đều bị tự động đóng + chấm 0★
# gần như ngay lập tức (before_request + cron 5 phút). Vì vậy bản tin này
# KHÔNG dùng khái niệm "quá hạn" mà dùng đúng bản chất: "không hoàn thành
# đúng hạn — 0★" cho việc đã bị tự động đóng, và liệt kê link để bấm vào
# xem/xử lý ngay thay vì chỉ đưa ra con số.

def _thu_tu_uu_tien(v: CongViec) -> int:
    return DoUuTien.THU_TU.get(v.do_uu_tien, 1)


def _dong_viec_kem_link(v: CongViec, ghi_chu: str = "") -> str:
    phu = f" — {ghi_chu}" if ghi_chu else ""
    return f"• [{v.ma}] {v.tieu_de}{phu}\n  {v.link}"


def _noi_dung_ban_tin_ca_nhan(nv: NguoiDung, buoi: str) -> str:
    """buoi: 'sang' (08h00, tổng kết hôm qua) hoặc 'chieu' (17h30, tổng kết
    hôm nay)."""
    base = current_app.config["BASE_URL"]
    hom_nay = ngay_vn_hien_tai()
    bay_gio = gio_vn_hien_tai()
    ngay_xet = hom_nay - timedelta(days=1) if buoi == "sang" else hom_nay
    tieu_de_ket_qua = "KẾT QUẢ HÔM QUA" if buoi == "sang" else "KẾT QUẢ HÔM NAY"
    tieu_de_ban_tin = "BẢN TIN SÁNG" if buoi == "sang" else "BẢN TIN CHIỀU"

    dong: list[str] = [
        f"🗞️ BRICON – {tieu_de_ban_tin} của {nv.ho_ten}",
        f"📅 {hom_nay:%d/%m/%Y}",
        "",
    ]

    # ---- Chấm công hôm nay ------------------------------------------------
    cc = ChamCong.query.filter_by(nguoi_dung_id=nv.id, ngay=hom_nay).first()
    dong.append("🟢 CHẤM CÔNG HÔM NAY")
    if cc and cc.gio_vao:
        if cc.di_tre:
            dong.append(f"🔴 Có mặt lúc {cc.gio_vao:%H:%M} — đi trễ {cc.so_phut_tre} phút")
        else:
            dong.append(f"✅ Có mặt lúc {cc.gio_vao:%H:%M} — đúng giờ")
    elif cc and cc.nghi_khong_phep:
        dong.append("🔴 Nghỉ không phép hôm nay")
    else:
        dong.append("⚪ Chưa chấm công vào — nhớ chấm công nhé!")
    dong.append("")

    # ---- Kết quả hôm qua / hôm nay -----------------------------------------
    dieu_kien = dieu_kien_viec_trong_ngay(ngay_xet)
    viec_ngay = CongViec.query.filter(
        CongViec.nguoi_nhan_id == nv.id, dieu_kien
    ).all()
    da_danh_gia = [v for v in viec_ngay if v.trang_thai == TrangThai.HOAN_THANH and v.so_sao_cuoi]
    khong_dung_han = [v for v in viec_ngay if v.trang_thai == TrangThai.HOAN_THANH and v.so_sao_cuoi == 0]
    cho_duyet_ngay = [v for v in viec_ngay if v.trang_thai == TrangThai.CHO_DUYET]
    tong_hoan_thanh = len(da_danh_gia) + len(khong_dung_han)

    dong.append(f"📊 {tieu_de_ket_qua}")
    dong.append(f"✅ Hoàn thành: {tong_hoan_thanh}/{len(viec_ngay)} công việc")
    dong.append(f"⭐ Đã đánh giá: {len(da_danh_gia)} việc")
    dong.append(f"🟡 Chờ đánh giá: {len(cho_duyet_ngay)} việc")
    dong.append(f"🔴 Không hoàn thành đúng hạn — 0★: {len(khong_dung_han)} việc")
    if khong_dung_han:
        dong.extend(_dong_viec_kem_link(v) for v in khong_dung_han)
    dong.append("")

    # ---- Việc quan trọng hôm nay --------------------------------------------
    dieu_kien_hom_nay = dieu_kien_viec_trong_ngay(hom_nay)
    mo_hom_nay = (
        CongViec.query.filter(
            CongViec.nguoi_nhan_id == nv.id,
            CongViec.trang_thai.in_(TrangThai.DANG_MO),
            dieu_kien_hom_nay,
        ).all()
    )
    mo_hom_nay.sort(key=lambda v: (_thu_tu_uu_tien(v), v.han is None, v.han))
    top3 = mo_hom_nay[:3]
    if top3:
        dong.append("🎯 VIỆC QUAN TRỌNG HÔM NAY")
        for i, v in enumerate(top3, 1):
            han_hien = f"trước {v.han:%H:%M}" if v.han else "không đặt hạn"
            dong.append(f"{i}. [{v.ma}] {v.tieu_de} — {han_hien}\n   {v.link}")
        dong.append("")

    # ---- Cần xử lý ngay: việc bị yêu cầu làm lại, chưa nộp lại --------------
    lam_lai_ds = CongViec.query.filter_by(
        nguoi_nhan_id=nv.id, trang_thai=TrangThai.LAM_LAI
    ).all()
    if lam_lai_ds:
        dong.append("⚠️ CẦN XỬ LÝ NGAY — bị yêu cầu làm lại")
        for v in lam_lai_ds:
            dg = v.danh_gia_theo_lan(v.lan_gui - 1)
            tu_luc = f" (từ {dg.tao_luc:%H:%M %d/%m})" if dg else ""
            dong.append(_dong_viec_kem_link(v, f"làm lại{tu_luc}"))
        dong.append("")

    # ---- Mức độ hiện tại (KPI tháng này) ------------------------------------
    bang_kpi = tinh_kpi(hom_nay.replace(day=1), hom_nay, nv.id)
    if bang_kpi:
        o = bang_kpi[0]
        if o["sao_tb"] is not None:
            dong.append(f"⭐ MỨC ĐỘ HIỆN TẠI (tháng này): {o['sao_tb']:.1f}/5 sao — {o['xep_loai']}")
            dong.append("")

    dong.append("💪 Mục tiêu: Hoàn thành 100%, không phát sinh việc bị 0★.")
    dong.append("")
    dong.append(f"👉 Xem việc của bạn: {base}/viec")
    dong.append(f"👉 Xem KPI: {base}/kpi")

    return "\n".join(dong)


def gui_ban_tin_sang() -> int:
    """08h00: gửi bản tin cá nhân buổi sáng cho từng nhân viên/quản lý (trừ
    Sếp/Admin) — tổng kết kết quả hôm qua + việc quan trọng hôm nay."""
    ds = _nhan_vien_khong_phai_admin_sep()
    for nv in ds:
        gui_cho_nhan_vien(nv, _noi_dung_ban_tin_ca_nhan(nv, "sang"))
    return len(ds)


def gui_ban_tin_chieu() -> int:
    """17h30: gửi bản tin cá nhân buổi chiều cho từng nhân viên/quản lý (trừ
    Sếp/Admin) — tổng kết kết quả hôm nay + việc còn tồn cần xử lý."""
    ds = _nhan_vien_khong_phai_admin_sep()
    for nv in ds:
        gui_cho_nhan_vien(nv, _noi_dung_ban_tin_ca_nhan(nv, "chieu"))
    return len(ds)


def bao_gui_doi_chung(viec: CongViec, so_file: int):
    """Báo cho người giao việc khi nhân viên vừa nộp đối chứng."""
    lan = f" (lần {viec.lan_gui})" if viec.lan_gui > 1 else ""
    nd = (
        f"✅ {viec.nguoi_nhan.ho_ten} đã gửi đối chứng{lan}\n\n"
        f"[{viec.ma}] {viec.tieu_de}\n"
        f"Số tệp đính kèm: {so_file}\n\n"
        f"Vào xem và đánh giá:\n{viec.link}"
    )
    gui_cho_nhan_vien(viec.nguoi_giao, nd, viec)


def bao_da_duyet(viec: CongViec, dg: DanhGia):
    sao = "⭐" * (dg.so_sao or 0)
    nhan_sao = MucSao.nhan(dg.so_sao)
    dong_danh_gia = f"Đánh giá: {sao} ({dg.so_sao}/5)"
    if nhan_sao:
        dong_danh_gia += f" - {nhan_sao}"
    nd = (
        f"🎉 Sếp đã xem và đánh giá công việc của bạn\n\n"
        f"[{viec.ma}] {viec.tieu_de}\n"
        f"{dong_danh_gia}\n"
    )
    if dg.ghi_chu:
        nd += f"Nhận xét: {dg.ghi_chu}\n"
    nd += f"\nXem lại:\n{viec.link}"
    gui_cho_nhan_vien(viec.nguoi_nhan, nd, viec)


def bao_danh_gia_lai(viec: CongViec, nguoi_sua: NguoiDung, sao_cu: int | None,
                      sao_moi: int, ly_do: str):
    """Báo cho nhân viên khi Admin/Sếp/Quản lý bộ phận sửa lại số sao của 1
    việc đã Hoàn thành (VD phát hiện vấn đề lúc bàn giao thực tế) — ảnh
    hưởng KPI của họ nên luôn báo, không âm thầm sửa."""
    sao_cu_hien = f"{sao_cu}★" if sao_cu is not None else "chưa có sao"
    nhan_sao_moi = MucSao.nhan(sao_moi)
    dong_nhan = f" - {nhan_sao_moi}" if nhan_sao_moi else ""
    nd = (
        f"✏️ Sếp vừa sửa lại đánh giá công việc của bạn\n\n"
        f"[{viec.ma}] {viec.tieu_de}\n"
        f"Đánh giá: {sao_cu_hien} → {sao_moi}★{dong_nhan}\n"
        f"Người sửa: {nguoi_sua.ho_ten}\n"
        f"Lý do: {ly_do}\n\n"
        f"Xem lại:\n{viec.link}"
    )
    gui_cho_nhan_vien(viec.nguoi_nhan, nd, viec)


def bao_lam_lai(viec: CongViec, dg: DanhGia):
    nd = (
        f"🔁 Công việc chưa đạt, cần làm lại\n\n"
        f"[{viec.ma}] {viec.tieu_de}\n"
        f"Người duyệt: {dg.nguoi_danh_gia.ho_ten}\n"
        f"Lý do: {dg.ghi_chu}\n\n"
        f"Làm lại rồi gửi đối chứng tại đây:\n{viec.link}"
    )
    gui_cho_nhan_vien(viec.nguoi_nhan, nd, viec)


# ---------------------------------------------------------------------------
# LƯU FILE
# ---------------------------------------------------------------------------
DUOI_ANH = {".jpg", ".jpeg", ".png", ".webp", ".heic", ".heif"}
DUOI_VIDEO = {".mp4", ".mov", ".webm", ".m4v", ".3gp"}
DUOI_AM = {".mp3", ".m4a", ".aac", ".ogg", ".wav", ".webm", ".amr"}
DUOI_FILE = {".pdf", ".doc", ".docx", ".xls", ".xlsx", ".txt", ".zip"}


def phan_loai(ten_file: str, mime: str = "") -> str | None:
    duoi = os.path.splitext(ten_file)[1].lower()
    mime = (mime or "").lower()
    if mime.startswith("image/") or duoi in DUOI_ANH:
        return LoaiDinhKem.ANH
    if mime.startswith("video/") or duoi in DUOI_VIDEO:
        return LoaiDinhKem.VIDEO
    if mime.startswith("audio/") or duoi in DUOI_AM:
        return LoaiDinhKem.GHI_AM
    if duoi in DUOI_FILE:
        return LoaiDinhKem.FILE
    return None


def _duong_dan_moi(ten_goc: str, thu_muc: str = "doi-chung") -> tuple[str, str]:
    """Trả về (đường dẫn tương đối — luôn dùng '/', đường dẫn tuyệt đối)."""
    hom_nay = ngay_vn_hien_tai()
    duoi = os.path.splitext(secure_filename(ten_goc))[1].lower() or ".bin"
    tuong_doi = f"{thu_muc}/{hom_nay:%Y/%m}/{uuid.uuid4().hex}{duoi}"
    tuyet_doi = os.path.join(current_app.config["UPLOAD_ROOT"], *tuong_doi.split("/"))
    os.makedirs(os.path.dirname(tuyet_doi), exist_ok=True)
    return tuong_doi, tuyet_doi


def luu_file(file_storage, thu_muc: str = "doi-chung") -> tuple[str, int]:
    tuong_doi, tuyet_doi = _duong_dan_moi(file_storage.filename, thu_muc)
    file_storage.save(tuyet_doi)
    return tuong_doi, os.path.getsize(tuyet_doi)


# ---------------------------------------------------------------------------
# ĐỊNH VỊ
# ---------------------------------------------------------------------------

def khoang_cach_m(lat1, lng1, lat2, lng2) -> float:
    """Haversine, trả về mét."""
    R = 6371000.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lng2 - lng1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * R * math.asin(math.sqrt(a))


def diem_gan_nhat(lat: float, lng: float):
    """Trả về (DiemChamCong|None, khoảng cách mét, có trong phạm vi không)."""
    diems = DiemChamCong.query.filter_by(dang_hoat_dong=True).all()
    if not diems:
        return None, None, True  # chưa cấu hình điểm nào thì không chặn
    gan_nhat, kc_min = None, None
    for d in diems:
        kc = khoang_cach_m(lat, lng, d.lat, d.lng)
        if kc_min is None or kc < kc_min:
            gan_nhat, kc_min = d, kc
    return gan_nhat, kc_min, kc_min <= gan_nhat.ban_kinh_m


def trung_toa_do_dang_ngo(nguoi_dung_id: int, lat: float, lng: float, so_ngay: int = 5) -> bool:
    """Cờ nghi ngờ: toạ độ trùng tới 6 chữ số thập phân nhiều ngày liên tiếp.

    GPS thật gần như không bao giờ ra kết quả giống hệt nhau; trùng khít
    thường là toạ độ do app fake GPS cắm cứng.
    """
    truoc = (
        ChamCong.query.filter(
            ChamCong.nguoi_dung_id == nguoi_dung_id,
            ChamCong.lat_vao.isnot(None),
            ChamCong.ngay >= ngay_vn_hien_tai() - timedelta(days=so_ngay * 2),
        )
        .order_by(ChamCong.ngay.desc())
        .limit(so_ngay)
        .all()
    )
    if len(truoc) < 2:
        return False
    return all(
        round(c.lat_vao, 6) == round(lat, 6) and round(c.lng_vao, 6) == round(lng, 6)
        for c in truoc[:2]
    )


# ---------------------------------------------------------------------------
# GIỜ LÀM VIỆC
# ---------------------------------------------------------------------------

def tinh_di_tre(gio: datetime, gio_chuan: tuple[int, int] | None = None,
                phut_tre_cho_phep: int | None = None) -> tuple[bool, int]:
    h, m = gio_chuan or current_app.config["GIO_VAO"]
    chuan = gio.replace(hour=h, minute=m, second=0, microsecond=0)
    phut = (phut_tre_cho_phep if phut_tre_cho_phep is not None
           else current_app.config["PHUT_TRE_CHO_PHEP"])
    han_mem = chuan + timedelta(minutes=phut)
    if gio <= han_mem:
        return False, 0
    return True, int((gio - chuan).total_seconds() // 60)


def tinh_ve_som(gio: datetime, gio_chuan: tuple[int, int] | None = None) -> tuple[bool, int]:
    h, m = gio_chuan or current_app.config["GIO_RA"]
    chuan = gio.replace(hour=h, minute=m, second=0, microsecond=0)
    if gio >= chuan:
        return False, 0
    return True, int((chuan - gio).total_seconds() // 60)


def tinh_lai_tre_som(cc: "ChamCong"):
    """Tính lại đi trễ/về sớm cho 1 bản ghi chấm công ĐÃ CÓ, theo đúng quy
    tắc hiện hành (kể cả nghỉ phép nửa ngày) — dùng để vá dữ liệu cũ chấm
    công trước khi có/đổi quy tắc này. Sửa thẳng lên object, không tự
    commit."""
    if cc.gio_vao:
        if co_nghi_phep_buoi(cc.nguoi_dung_id, cc.ngay, BuoiNghi.SANG):
            cc.di_tre, cc.so_phut_tre = tinh_di_tre(
                cc.gio_vao, current_app.config["GIO_BAT_DAU_CHIEU"],
                current_app.config["PHUT_TRE_CHO_PHEP_NUA_NGAY"])
        elif co_nghi_phep_buoi(cc.nguoi_dung_id, cc.ngay, BuoiNghi.CHIEU):
            cc.di_tre, cc.so_phut_tre = tinh_di_tre(
                cc.gio_vao, current_app.config["GIO_VAO"],
                current_app.config["PHUT_TRE_CHO_PHEP_NUA_NGAY"])
        else:
            cc.di_tre, cc.so_phut_tre = tinh_di_tre(cc.gio_vao)

    if cc.gio_ra:
        if co_nghi_phep_buoi(cc.nguoi_dung_id, cc.ngay, BuoiNghi.CHIEU):
            cc.ve_som, cc.so_phut_som = tinh_ve_som(
                cc.gio_ra, current_app.config["GIO_KET_THUC_SANG"])
        else:
            cc.ve_som, cc.so_phut_som = tinh_ve_som(cc.gio_ra)


def gio_cong_trong_ngay(cc: "ChamCong") -> float:
    """Tính số công (0 / 0.5 / 1) của 1 lần chấm công — CHỈ dựa theo giờ
    vào/ra thực tế cắt theo ranh giới trưa, không liên quan gì tới nghỉ
    phép (nghỉ phép không được cộng thêm công, đúng theo yêu cầu "công chỉ
    tính khi chấm công tại địa điểm").

    Có mặt buổi sáng: chấm vào lúc hoặc trước giờ bắt đầu buổi chiều.
    Có mặt buổi chiều: đã chấm ra, lúc hoặc sau giờ kết thúc buổi sáng.
    """
    if not cc.gio_vao:
        return 0.0
    h_sang, m_sang = current_app.config["GIO_KET_THUC_SANG"]
    h_chieu, m_chieu = current_app.config["GIO_BAT_DAU_CHIEU"]
    moc_ket_thuc_sang = cc.gio_vao.replace(hour=h_sang, minute=m_sang, second=0, microsecond=0)
    moc_bat_dau_chieu = cc.gio_vao.replace(hour=h_chieu, minute=m_chieu, second=0, microsecond=0)

    co_lam_sang = cc.gio_vao <= moc_bat_dau_chieu
    co_lam_chieu = bool(cc.gio_ra) and cc.gio_ra >= moc_ket_thuc_sang
    return (0.5 if co_lam_sang else 0.0) + (0.5 if co_lam_chieu else 0.0)


# ---------------------------------------------------------------------------
# MÃ CÔNG VIỆC
# ---------------------------------------------------------------------------

def gan_ma(viec: CongViec):
    """Mã sinh từ khoá chính Postgres -> không bao giờ trùng, không cần đếm."""
    viec.ma = f"V{viec.id + current_app.config['TASK_CODE_OFFSET']}"


# ---------------------------------------------------------------------------
# KPI — thang 0-5 sao theo tiêu chí thái độ làm việc
# ---------------------------------------------------------------------------
NHAN_SAO = {
    0: "Không làm",
    1: "Làm cho có",
    2: "Có làm nhưng cần nhắc nhở",
    3: "Làm đầy đủ, có trách nhiệm",
    4: "Có tâm, nhiệt huyết",
    5: "Chủ động, sáng tạo, vượt mong đợi",
}

GIA_HAN_LAM_LAI = timedelta(hours=3)  # cộng thêm kể từ lúc sếp bấm "Yêu cầu làm lại"


def _han_hieu_luc(v: CongViec) -> datetime | None:
    """Hạn dùng để xét quá hạn khi tính KPI.

    Mặc định là hạn gốc (v.han). Mỗi lần sếp yêu cầu làm lại trong lúc việc
    đã quá hạn gốc (đã nộp, đang chờ duyệt, nhưng trễ), nhân viên được cộng
    thêm GIA_HAN_LAM_LAI kể từ đúng lúc sếp bấm "Yêu cầu làm lại" — nếu vẫn
    tính theo hạn cũ thì coi như họ không còn chút thời gian nào để sửa lại.
    Nhiều lần làm lại liên tiếp thì cộng dồn theo từng lần.
    """
    if not v.han:
        return None
    han = v.han
    for dg in v.danh_gia:
        if dg.ket_qua == "lam_lai" and dg.tao_luc > han:
            han = dg.tao_luc + GIA_HAN_LAM_LAI
    return han


def nhan_theo_sao(sao_tb: float | None) -> str:
    if sao_tb is None:
        return "— Chưa có dữ liệu"
    muc = min(5, max(0, math.floor(sao_tb + 0.5)))
    return f"{muc}★ {NHAN_SAO[muc]}"


def tinh_kpi(tu_ngay: date, den_ngay: date, nguoi_dung_id: int | None = None) -> list[dict]:
    """Tính KPI trong khoảng ngày [tu_ngay, den_ngay].

    - Việc đã Hoàn thành: tính theo ngày hoàn thành (hoan_thanh_luc) nằm
      trong khoảng, dùng đúng số sao sếp đã chấm.
    - Việc chưa xong (mới/đang làm/làm lại/chờ duyệt): tính theo hạn (han)
      nằm trong khoảng — để bắt được cả việc quá hạn mà chưa nộp gì.
    - Quá hạn hiệu lực mà chưa nộp gì lại (moi/dang_lam/lam_lai) -> tự động 0★.
    - Đang Chờ duyệt (đã nộp, sếp chưa xem) -> không tự chấm dù có quá hạn
      hay không, để sếp vào đánh giá tay.
    """
    dau = datetime.combine(tu_ngay, datetime.min.time())
    cuoi = datetime.combine(den_ngay, datetime.max.time())
    bay_gio = gio_vn_hien_tai()

    q_xong = CongViec.query.join(
        NguoiDung, CongViec.nguoi_nhan_id == NguoiDung.id
    ).filter(
        NguoiDung.vai_tro.notin_((VaiTro.ADMIN, VaiTro.SEP)),
        CongViec.trang_thai == TrangThai.HOAN_THANH,
        CongViec.hoan_thanh_luc.isnot(None),
        CongViec.hoan_thanh_luc >= dau,
        CongViec.hoan_thanh_luc <= cuoi,
    )
    q_con_lai = CongViec.query.join(
        NguoiDung, CongViec.nguoi_nhan_id == NguoiDung.id
    ).filter(
        NguoiDung.vai_tro.notin_((VaiTro.ADMIN, VaiTro.SEP)),
        CongViec.han.isnot(None),
        CongViec.han >= dau,
        CongViec.han <= cuoi,
        CongViec.trang_thai.notin_((TrangThai.HOAN_THANH, TrangThai.HUY)),
    )
    if nguoi_dung_id:
        q_xong = q_xong.filter(CongViec.nguoi_nhan_id == nguoi_dung_id)
        q_con_lai = q_con_lai.filter(CongViec.nguoi_nhan_id == nguoi_dung_id)

    viecs = {v.id: v for v in q_xong.all()}
    for v in q_con_lai.all():
        viecs.setdefault(v.id, v)

    bang: dict[int, dict] = {}
    for v in viecs.values():
        o = bang.setdefault(
            v.nguoi_nhan_id,
            {
                "nguoi_dung_id": v.nguoi_nhan_id,
                "ho_ten": v.nguoi_nhan.ho_ten,
                "ma_dinh_danh": v.nguoi_nhan.ma_dinh_danh,
                "tong_viec": 0, "da_danh_gia": 0, "cho_duyet": 0, "dang_lam": 0,
                "tong_sao": 0,
            },
        )
        o["tong_viec"] += 1

        if v.trang_thai == TrangThai.HOAN_THANH:
            if v.so_sao_cuoi is not None:
                o["da_danh_gia"] += 1
                o["tong_sao"] += v.so_sao_cuoi
            continue

        if v.trang_thai == TrangThai.CHO_DUYET:
            o["cho_duyet"] += 1
            continue

        # còn lại: moi / dang_lam / lam_lai
        han = _han_hieu_luc(v)
        if han and bay_gio > han:
            o["da_danh_gia"] += 1
            o["tong_sao"] += 0
        else:
            o["dang_lam"] += 1

    ds = []
    for o in bang.values():
        o["sao_tb"] = round(o["tong_sao"] / o["da_danh_gia"], 2) if o["da_danh_gia"] else None
        o["xep_loai"] = nhan_theo_sao(o["sao_tb"])
        ds.append(o)
    ds.sort(key=lambda x: (x["sao_tb"] is None, -(x["sao_tb"] or 0)))
    for i, o in enumerate(ds, 1):
        o["hang"] = i
    return ds


# ---------------------------------------------------------------------------
# TỰ ĐỘNG ĐÓNG VIỆC QUÁ HẠN
# ---------------------------------------------------------------------------

def bao_tu_dong_dong(viec: CongViec):
    nd_nv = (
        f"🔒 Công việc đã tự động đóng do quá hạn không nộp đối chứng\n\n"
        f"[{viec.ma}] {viec.tieu_de}\n"
        f"Kết quả: 0★ — Không làm\n\n"
        f"Việc đã bị khoá lại, chỉ Admin hoặc Ban giám đốc mới mở lại được.\n{viec.link}"
    )
    gui_cho_nhan_vien(viec.nguoi_nhan, nd_nv, viec)

    nd_ql = (
        f"🔴 Việc tự động đóng 0★ — quá hạn không làm\n\n"
        f"[{viec.ma}] {viec.tieu_de}\n"
        f"Người thực hiện: {viec.nguoi_nhan.ho_ten} ({viec.nguoi_nhan.ma_dinh_danh})\n"
        f"Người giao: {viec.nguoi_giao.ho_ten}\n"
        f"Hạn: {_han_str(viec)}\n\n"
        f"{viec.link}"
    )
    gui_nhom_ql(nd_ql, viec)


def dong_cac_viec_qua_han() -> list[CongViec]:
    """Tự động đóng + chấm 0★ các việc đã quá hạn hiệu lực mà chưa nộp đối
    chứng (mới/đang làm/làm lại — không đụng tới việc đang Chờ duyệt, luôn
    để sếp tự chấm). Gọi hàm này ở nhiều điểm (mỗi lần tải trang liên quan
    tới công việc, và cả qua cron) để việc được đóng gần như ngay khi quá hạn.
    """
    bay_gio = gio_vn_hien_tai()
    ung_vien = CongViec.query.filter(
        CongViec.trang_thai.in_(TrangThai.CHUA_XONG),
        CongViec.han.isnot(None),
        CongViec.han < bay_gio,
    ).all()

    da_dong = []
    for v in ung_vien:
        han = _han_hieu_luc(v)
        if han and bay_gio > han:
            v.trang_thai = TrangThai.HOAN_THANH
            v.so_sao_cuoi = 0
            v.hoan_thanh_luc = bay_gio
            db.session.add(DanhGia(
                cong_viec_id=v.id, nguoi_danh_gia_id=v.nguoi_giao_id,
                lan_gui=v.lan_gui, ket_qua="dat", so_sao=0,
                ghi_chu="Hệ thống tự động đóng: quá hạn mà không nộp đối chứng.",
            ))
            da_dong.append(v)

    if da_dong:
        db.session.commit()
        for v in da_dong:
            bao_tu_dong_dong(v)
        db.session.commit()  # lưu log zalo

    return da_dong


def go_lien_ket_log_zalo_cho_viec(viec: CongViec):
    """Gỡ liên kết log Zalo khỏi 1 công việc trước khi xoá vĩnh viễn công
    việc đó — Postgres chặn xoá nếu còn log tham chiếu qua khoá ngoại
    cong_viec_id. Giữ nguyên log để tra cứu, chỉ gỡ liên kết."""
    for lz in LogZalo.query.filter_by(cong_viec_id=viec.id).all():
        lz.cong_viec_id = None


def xoa_file_dinh_kem(viec: CongViec):
    """Xoá vật lý các tệp đối chứng của 1 công việc trên ổ đĩa, gọi trước khi
    xoá bản ghi CongViec để không để lại file mồ côi."""
    for d in viec.dinh_kem:
        duong_dan = os.path.join(current_app.config["UPLOAD_ROOT"], *d.duong_dan.split("/"))
        try:
            os.remove(duong_dan)
        except OSError:
            pass


def xoa_toan_bo_du_lieu_nhan_vien(nd: NguoiDung, admin_thuc_hien: NguoiDung):
    """Xoá vĩnh viễn 1 nhân viên và TOÀN BỘ dữ liệu gắn với chính họ —
    không thể khôi phục. Chỉ gọi từ route đã tự kiểm tra quyền Admin thuần.

    - Việc họ NHẬN: xoá hẳn, kèm file đính kèm vật lý và đánh giá đi theo
      (đã cascade sẵn qua quan hệ CongViec.dinh_kem / CongViec.danh_gia).
    - Việc họ từng GIAO cho người khác: KHÔNG xoá (sẽ mất lịch sử của
      người nhận việc, không liên quan gì tới người bị xoá) — chỉ đổi
      người giao thành admin đang thực hiện xoá, kèm ghi chú.
    - Đánh giá họ từng chấm cho việc người khác: giữ lại, đổi người
      đánh giá thành admin đang thực hiện xoá.
    - Chấm công, xin nghỉ (kèm ảnh minh chứng vật lý): xoá hẳn — dữ liệu
      của riêng họ.
    - Log Zalo gắn với họ: gỡ liên kết (giữ log để tra cứu, không xoá).
    """
    from models import XinNghi

    for v in CongViec.query.filter_by(nguoi_nhan_id=nd.id).all():
        xoa_file_dinh_kem(v)
        go_lien_ket_log_zalo_cho_viec(v)
        db.session.delete(v)

    for v in CongViec.query.filter_by(nguoi_giao_id=nd.id).all():
        v.nguoi_giao_id = admin_thuc_hien.id
        v.mo_ta = (v.mo_ta or "") + (
            f"\n\n[Người giao gốc {nd.ho_ten} ({nd.ma_dinh_danh}) đã bị xoá tài "
            f"khoản — chuyển người giao sang {admin_thuc_hien.ho_ten}]"
        )

    for dg in DanhGia.query.filter_by(nguoi_danh_gia_id=nd.id).all():
        dg.nguoi_danh_gia_id = admin_thuc_hien.id

    for cc in ChamCong.query.filter_by(nguoi_dung_id=nd.id).all():
        db.session.delete(cc)

    for x in XinNghi.query.filter_by(nguoi_dung_id=nd.id).all():
        duong_dan = os.path.join(current_app.config["UPLOAD_ROOT"], *x.anh_minh_chung.split("/"))
        try:
            os.remove(duong_dan)
        except OSError:
            pass
        db.session.delete(x)

    for lz in LogZalo.query.filter_by(nguoi_dung_id=nd.id).all():
        lz.nguoi_dung_id = None

    db.session.delete(nd)


# ---------------------------------------------------------------------------
# XUẤT EXCEL
# ---------------------------------------------------------------------------
_NEN_TIEU_DE = PatternFill("solid", fgColor="111111")
_CHU_TIEU_DE = Font(bold=True, color="FFFFFF")
_GIUA = Alignment(horizontal="center")


def _ke_tieu_de(ws, do_rong: list[int]):
    for cell in ws[1]:
        cell.font = _CHU_TIEU_DE
        cell.fill = _NEN_TIEU_DE
        cell.alignment = _GIUA
    for i, w in enumerate(do_rong, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w


def xuat_excel_bang_cong(thang: str, ban_ghi: list, don_nghi: list, tong: list) -> BytesIO:
    """Xuất bảng công ra file Excel — 3 sheet: Tổng hợp, Chi tiết chấm công,
    và Nghỉ phép trong tháng."""
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Tổng hợp"
    ws1.append(["Mã NV", "Họ tên", "Ngày công", "Trễ", "Về sớm", "Nghỉ phép", "Không phép"])
    for o in tong:
        ws1.append([o["ma"], o["ho_ten"], o["so_ngay"], o["tre"], o["som"], o["nghi_phep"], o["khong_phep"]])
    _ke_tieu_de(ws1, [10, 26, 11, 8, 10, 11, 11])

    ws2 = wb.create_sheet("Chi tiết chấm công")
    ws2.append(["Ngày", "Mã NV", "Họ tên", "Giờ vào", "Giờ ra", "Điểm chấm công",
               "Đi trễ (phút)", "Về sớm (phút)", "Nghỉ không phép", "Nghi fake GPS"])
    for c in ban_ghi:
        ws2.append([
            c.ngay.strftime("%d/%m/%Y"),
            c.nguoi_dung.ma_dinh_danh,
            c.nguoi_dung.ho_ten,
            c.gio_vao.strftime("%H:%M") if c.gio_vao else "",
            c.gio_ra.strftime("%H:%M") if c.gio_ra else "",
            c.diem_vao.ten if c.diem_vao else "",
            c.so_phut_tre if c.di_tre else "",
            c.so_phut_som if c.ve_som else "",
            "Có" if c.nghi_khong_phep else "",
            "Có" if c.nghi_ngo else "",
        ])
    _ke_tieu_de(ws2, [12, 10, 26, 10, 10, 22, 14, 14, 15, 14])

    ws3 = wb.create_sheet("Nghỉ phép")
    ws3.append(["Ngày", "Mã NV", "Họ tên", "Buổi", "Ghi chú"])
    for x in don_nghi:
        ws3.append([
            x.ngay.strftime("%d/%m/%Y"),
            x.nguoi_dung.ma_dinh_danh,
            x.nguoi_dung.ho_ten,
            x.ten_buoi,
            x.ghi_chu or "",
        ])
    _ke_tieu_de(ws3, [12, 10, 26, 12, 30])

    dem = BytesIO()
    wb.save(dem)
    dem.seek(0)
    return dem


def xuat_excel_tro_ly_su_dung(thang: str, chi_tiet: list, tong: list) -> BytesIO:
    """Xuất thống kê sử dụng Trợ lý AI ra file Excel — 2 sheet: Tổng hợp
    theo người (cả tháng) và Chi tiết theo từng ngày."""
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Tổng hợp"
    ws1.append(["Mã NV", "Họ tên", "Vai trò", "Tổng số câu hỏi", "Tổng số token"])
    for o in tong:
        ws1.append([o["ma"], o["ho_ten"], o["vai_tro"], o["so_cau_hoi"], o["so_token"]])
    _ke_tieu_de(ws1, [10, 26, 18, 16, 16])

    ws2 = wb.create_sheet("Chi tiết theo ngày")
    ws2.append(["Ngày", "Mã NV", "Họ tên", "Vai trò", "Số câu hỏi", "Số token"])
    for su_dung in chi_tiet:
        ws2.append([
            su_dung.ngay.strftime("%d/%m/%Y"),
            su_dung.nguoi_dung.ma_dinh_danh,
            su_dung.nguoi_dung.ho_ten,
            su_dung.nguoi_dung.ten_vai_tro,
            su_dung.so_cau_hoi,
            su_dung.so_token,
        ])
    _ke_tieu_de(ws2, [12, 10, 26, 18, 12, 12])

    dem = BytesIO()
    wb.save(dem)
    dem.seek(0)
    return dem


# ---------------------------------------------------------------------------
# CẤU HÌNH HỆ THỐNG (key-value đơn giản)
# ---------------------------------------------------------------------------

def lay_cai_dat(khoa: str, mac_dinh: str | None = None) -> str | None:
    from models import CaiDat
    cd = db.session.get(CaiDat, khoa)
    return cd.gia_tri if cd and cd.gia_tri else mac_dinh


def dat_cai_dat(khoa: str, gia_tri: str):
    from models import CaiDat
    cd = db.session.get(CaiDat, khoa)
    if not cd:
        cd = CaiDat(khoa=khoa)
        db.session.add(cd)
    cd.gia_tri = gia_tri


# ---------------------------------------------------------------------------
# DỌN DẸP DỮ LIỆU CŨ — tách "--- Kết quả lần N ---" khỏi mô tả
# ---------------------------------------------------------------------------
_MAU_GHI_CHU_CU = re.compile(r"\n\n--- Kết quả lần (\d+) ---\n")


def tach_ghi_chu_nop_cu() -> tuple[int, int]:
    """Dọn dữ liệu cũ: trước đây khi nhân viên gửi đối chứng kèm ghi chú,
    hệ thống nối thẳng ghi chú đó vào CongViec.mo_ta (đánh dấu bằng
    "--- Kết quả lần N ---"), khiến mô tả gốc bị lẫn với ghi chú của từng
    lần nộp và ghi chú không hiện đúng chỗ "Lần nộp N" nữa.

    Hàm này quét mọi CongViec còn dính mẫu trên, tách từng đoạn ghi chú
    ra thành 1 dòng GhiChuNop riêng (đúng lan_gui), rồi cắt mô tả về lại
    đúng phần gốc (trước đoạn "--- Kết quả lần" đầu tiên).

    Chạy lại nhiều lần vẫn an toàn: đã tách rồi (đã có GhiChuNop cho đúng
    cong_viec_id + lan_gui) thì bỏ qua, không tạo trùng.

    Trả về (số CongViec đã xử lý, số GhiChuNop mới tạo).
    """
    so_viec_xu_ly = 0
    so_ghi_chu_tao = 0

    viecs = CongViec.query.filter(
        CongViec.mo_ta.ilike("%--- Kết quả lần%")
    ).all()

    for v in viecs:
        cac_phan = _MAU_GHI_CHU_CU.split(v.mo_ta or "")
        if len(cac_phan) < 3:
            continue  # có chữ "--- Kết quả lần" nhưng không đúng mẫu -> bỏ qua, an toàn

        mo_ta_goc = cac_phan[0].rstrip()
        da_dong_nao = False

        i = 1
        while i < len(cac_phan) - 1:
            lan = int(cac_phan[i])
            noi_dung = cac_phan[i + 1].rstrip("\n")
            i += 2

            if not noi_dung:
                continue
            da_ton_tai = GhiChuNop.query.filter_by(
                cong_viec_id=v.id, lan_gui=lan
            ).first()
            if da_ton_tai:
                continue

            db.session.add(GhiChuNop(
                cong_viec_id=v.id, lan_gui=lan, noi_dung=noi_dung,
                tao_luc=v.gui_doi_chung_luc or v.tao_luc,
            ))
            so_ghi_chu_tao += 1
            da_dong_nao = True

        if da_dong_nao:
            v.mo_ta = mo_ta_goc or None
            so_viec_xu_ly += 1

    db.session.commit()
    return so_viec_xu_ly, so_ghi_chu_tao